"""Integration tests for the Phase 4 ``/reports/expenses-excel`` endpoint.

Hits the real ASGI app with admin / contributor JWTs and exercises:

* 200 happy path with valid `.xlsx` body + correct Content-Type +
  RFC 5987 dual-form Content-Disposition
* `include_pending=true` reflects in the workbook annotation
* `from_date` / `to_date` reflect in the workbook annotation +
  filter the rows
* `from_date > to_date` → 400
* unknown `job_id` → 404
* contributor token → 403
* no token → 401
* CJK job name → Content-Disposition carries both ASCII fallback and
  UTF-8 (RFC 5987) forms

Workbook contents are covered in depth by
``tests/test_excel_export.py``; these tests stay focused on the HTTP
contract (auth, status codes, headers, wire shape).
"""

from __future__ import annotations

import datetime as _datetime
import uuid
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.models.expense import (
    Expense,
    ExpenseType,
    PaymentMethod,
    ReceiptStatus,
    ReviewStatus,
)
from app.models.job import Job, JobStatus

_XLSX_MIME = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


async def _mk_job(
    db, admin, *, name: str = "Job", code: str | None = None
) -> Job:
    j = Job(
        job_id=uuid.uuid4(),
        job_name=name,
        job_code=code,
        status=JobStatus.active,
        created_by=admin.user_id,
    )
    db.add(j)
    await db.flush()
    return j


async def _mk_expense(
    db,
    *,
    job: Job,
    admin,
    amount_inc_gst: Decimal = Decimal("110.00"),
    review_status: ReviewStatus = ReviewStatus.reviewed,
    expense_date: _datetime.date | None = None,
) -> Expense:
    e = Expense(
        expense_id=uuid.uuid4(),
        job_id=job.job_id,
        entered_by_user_id=admin.user_id,
        expense_type=ExpenseType.supplier_expense,
        amount_inc_gst=amount_inc_gst,
        payment_method=PaymentMethod.transfer,
        expense_date=expense_date or _datetime.date.today(),
        review_status=review_status,
        receipt_status=ReceiptStatus.no_receipt,
    )
    db.add(e)
    await db.flush()
    return e


# ---------------------------------------------------------------------------
# Auth + HTTP contract
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_get_expenses_excel_401_no_token(client):
    r = await client.get("/reports/expenses-excel")
    assert r.status_code == 401


@pytest.mark.asyncio
async def test_get_expenses_excel_403_contributor(
    client, db_session, seeded_admin, contributor_token
):
    job = await _mk_job(db_session, seeded_admin, name="ContribForbid")
    await _mk_expense(db_session, job=job, admin=seeded_admin)
    r = await client.get(
        "/reports/expenses-excel",
        headers={"Authorization": f"Bearer {contributor_token}"},
    )
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_get_expenses_excel_200_admin_returns_valid_xlsx(
    client, db_session, seeded_admin, admin_token
):
    job = await _mk_job(db_session, seeded_admin, name="200OK")
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, amount_inc_gst=Decimal("550.00")
    )
    r = await client.get(
        "/reports/expenses-excel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    assert r.headers["content-type"] == _XLSX_MIME
    assert "attachment" in r.headers["content-disposition"]
    # Body parses as a valid xlsx with both expected sheets.
    wb = load_workbook(BytesIO(r.content))
    assert "All Expenses" in wb.sheetnames
    assert "200OK" in wb.sheetnames


@pytest.mark.asyncio
async def test_get_expenses_excel_400_on_date_range_inverted(
    client, admin_token
):
    r = await client.get(
        "/reports/expenses-excel?from_date=2026-12-31&to_date=2026-01-01",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 400


@pytest.mark.asyncio
async def test_get_expenses_excel_400_on_malformed_date(client, admin_token):
    r = await client.get(
        "/reports/expenses-excel?from_date=not-a-date",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    # FastAPI rejects malformed query params with 422 by default; that's
    # acceptable too. Either 400 or 422 means the bad input was caught.
    assert r.status_code in (400, 422)


@pytest.mark.asyncio
async def test_get_expenses_excel_404_unknown_job(client, admin_token):
    bogus = uuid.uuid4()
    r = await client.get(
        f"/reports/expenses-excel?job_id={bogus}",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 404


# ---------------------------------------------------------------------------
# Query-param semantics reflected in workbook
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_get_expenses_excel_default_excludes_pending(
    client, db_session, seeded_admin, admin_token
):
    job = await _mk_job(db_session, seeded_admin, name="DefaultExcl")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        review_status=ReviewStatus.pending,
    )
    r = await client.get(
        "/reports/expenses-excel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "DefaultExcl" not in wb.sheetnames
    ws = wb["All Expenses"]
    assert "reviewed expenses only" in ws.cell(row=2, column=1).value


@pytest.mark.asyncio
async def test_get_expenses_excel_include_pending(
    client, db_session, seeded_admin, admin_token
):
    job = await _mk_job(db_session, seeded_admin, name="WithPending")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        review_status=ReviewStatus.pending,
    )
    r = await client.get(
        "/reports/expenses-excel?include_pending=true",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    assert "WithPending" in wb.sheetnames
    ws = wb["All Expenses"]
    assert "reviewed + pending" in ws.cell(row=2, column=1).value


@pytest.mark.asyncio
async def test_get_expenses_excel_date_range_reflected(
    client, db_session, seeded_admin, admin_token
):
    job = await _mk_job(db_session, seeded_admin, name="DateRange")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        expense_date=_datetime.date(2026, 5, 5),
        amount_inc_gst=Decimal("110.00"),
    )
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        expense_date=_datetime.date(2026, 1, 1),
        amount_inc_gst=Decimal("330.00"),
    )
    r = await client.get(
        "/reports/expenses-excel?from_date=2026-05-01&to_date=2026-05-31",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb["All Expenses"]
    # Period label reflects the requested range.
    period = ws.cell(row=3, column=1).value
    assert "2026-05-01" in period
    assert "2026-05-31" in period
    # Only the in-range row appears.
    # Row 6 first data, row 7 totals.
    assert ws.cell(row=7, column=1).value == "Totals"
    assert ws.cell(row=7, column=7).value == Decimal("110.00")


# ---------------------------------------------------------------------------
# Content-Disposition (RFC 5987 dual-form, CJK safety)
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_content_disposition_default_today_stamp(
    client, db_session, seeded_admin, admin_token
):
    r = await client.get(
        "/reports/expenses-excel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    cd = r.headers["content-disposition"]
    today = _datetime.date.today().isoformat()
    assert today in cd
    assert "sitetracker-export-" in cd


@pytest.mark.asyncio
async def test_content_disposition_date_range_stamp(client, admin_token):
    r = await client.get(
        "/reports/expenses-excel?from_date=2025-07-01&to_date=2026-06-30",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    cd = r.headers["content-disposition"]
    assert "sitetracker-export-2025-07-01-to-2026-06-30.xlsx" in cd


@pytest.mark.asyncio
async def test_content_disposition_cjk_job_dual_form(
    client, db_session, seeded_admin, admin_token
):
    """晶晶 job: ASCII fallback uses job_id prefix; UTF-8 form preserves CJK."""
    job = await _mk_job(db_session, seeded_admin, name="晶晶")
    await _mk_expense(db_session, job=job, admin=seeded_admin)
    r = await client.get(
        f"/reports/expenses-excel?job_id={job.job_id}",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    cd = r.headers["content-disposition"]
    # ASCII fallback can't carry CJK — must use the job_id hex prefix.
    job_id_prefix = str(job.job_id).replace("-", "")[:8]
    assert job_id_prefix in cd
    # RFC 5987 form: filename*=UTF-8'' followed by percent-encoded CJK.
    assert "filename*=UTF-8''" in cd
    # 晶 = U+6676 → UTF-8 bytes E6 99 B6 → percent-encoded %E6%99%B6
    assert "%E6%99%B6" in cd


@pytest.mark.asyncio
async def test_content_disposition_is_attachment(client, admin_token):
    r = await client.get(
        "/reports/expenses-excel",
        headers={"Authorization": f"Bearer {admin_token}"},
    )
    assert r.headers["content-disposition"].lower().startswith("attachment")


# ---------------------------------------------------------------------------
# Phase 4 Batch 2 hardening — CORS exposure of Content-Disposition
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_cors_exposes_content_disposition(client, admin_token):
    """CORS must expose ``Content-Disposition`` to cross-origin JS.

    Without ``Access-Control-Expose-Headers: Content-Disposition``
    on the CORS response, browsers silently strip the header from
    cross-origin responses and the admin's filename-parsing path
    falls through to the hard-coded fallback ``sitetracker-export.xlsx``
    — losing CJK job names and the date-stamp. Live E2E in Batch 2
    surfaced this; this regression test guards the fix
    (``expose_headers=["Content-Disposition"]`` in ``app/main.py``'s
    CORS config) so it can't quietly regress.

    We exercise the path by sending an ``Origin`` header on the GET
    request, which is what triggers FastAPI's CORSMiddleware to emit
    the CORS response headers.
    """
    r = await client.get(
        "/reports/expenses-excel",
        headers={
            "Authorization": f"Bearer {admin_token}",
            "Origin": "http://localhost:5173",
        },
    )
    assert r.status_code == 200, r.text
    # Both headers must be present.
    expose = r.headers.get("access-control-expose-headers", "")
    assert "Content-Disposition" in expose, (
        f"CORS Access-Control-Expose-Headers missing Content-Disposition; "
        f"got: {expose!r}"
    )
    # And the header itself still rides on the response.
    cd = r.headers.get("content-disposition", "")
    assert cd.lower().startswith("attachment"), (
        f"Content-Disposition missing or not an attachment: {cd!r}"
    )
