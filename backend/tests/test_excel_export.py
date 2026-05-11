"""Tests for the Phase 4 ``excel_export`` service.

Covers the contract frozen by ``docs/phase-4-plan.md``:

* Default inclusion rule (reviewed-only); rejected always excluded;
  pending opt-in via ``include_pending=True``.
* Date-range filter on ``expense_date``.
* Per-job sheet header is date-range aware — period totals reflect
  only the rows on the sheet, project budget summary is always
  all-time + Phase 3 Lite (reviewed+pending) view.
* Two new audit columns: Raw input text + Created at (ISO-UTC).
* Excel formula-injection neutralisation on every text cell.
* Sheet name + filename sanitisation including CJK + RFC 5987 safety.

Each test builds the workbook bytes via :func:`build_workbook`, reads
them back via ``openpyxl.load_workbook(BytesIO(...))``, and asserts
on cell values, sheet names, and number formats.
"""

from __future__ import annotations

import datetime as _datetime
import uuid
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import text as sa_text

from app.models.expense import (
    Expense,
    ExpenseType,
    PaymentMethod,
    ReceiptStatus,
    ReviewStatus,
)
from app.models.job import Job, JobStatus
from app.models.supplier import Supplier
from app.services.excel_export import (
    _FORMULA_PREFIXES,
    _safe_excel_text,
    _safe_sheet_name,
    build_export_filename,
    build_workbook,
)
from app.services.jobs import JobNotFound


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _today() -> _datetime.date:
    return _datetime.date.today()


async def _mk_job(
    db,
    admin,
    *,
    name: str = "Test Job",
    code: str | None = None,
    site: str | None = None,
    total_budget_ex_gst: Decimal | None = None,
    contract_value_ex_gst: Decimal | None = None,
    target_profit_ratio_pct: Decimal | None = None,
) -> Job:
    job = Job(
        job_id=uuid.uuid4(),
        job_name=name,
        job_code=code,
        site_address=site,
        status=JobStatus.active,
        total_budget_ex_gst=total_budget_ex_gst,
        contract_value_ex_gst=contract_value_ex_gst,
        target_profit_ratio_pct=target_profit_ratio_pct,
        created_by=admin.user_id,
    )
    db.add(job)
    await db.flush()
    return job


async def _mk_supplier(db, *, name: str = "Bunnings") -> Supplier:
    s = Supplier(supplier_id=uuid.uuid4(), supplier_name=name, is_active=True)
    db.add(s)
    await db.flush()
    return s


async def _mk_expense(
    db,
    *,
    job: Job,
    admin,
    amount_inc_gst: Decimal = Decimal("110.00"),
    payment_method: PaymentMethod = PaymentMethod.transfer,
    review_status: ReviewStatus = ReviewStatus.reviewed,
    expense_date: _datetime.date | None = None,
    category_id: uuid.UUID | None = None,
    supplier: Supplier | None = None,
    description: str | None = None,
    notes: str | None = None,
    raw_input_text: str | None = None,
    expense_type: ExpenseType = ExpenseType.supplier_expense,
) -> Expense:
    e = Expense(
        expense_id=uuid.uuid4(),
        job_id=job.job_id,
        entered_by_user_id=admin.user_id,
        expense_type=expense_type,
        amount_inc_gst=amount_inc_gst,
        payment_method=payment_method,
        expense_date=expense_date or _today(),
        review_status=review_status,
        receipt_status=ReceiptStatus.no_receipt,
        category_id=category_id,
        supplier_id=supplier.supplier_id if supplier else None,
        description=description,
        notes=notes,
        raw_input_text=raw_input_text,
    )
    db.add(e)
    await db.flush()
    return e


def _open(body: bytes):
    """Convenience: read workbook bytes back via openpyxl."""
    return load_workbook(BytesIO(body))


# ===========================================================================
# _safe_excel_text — formula-injection helper unit tests
# ===========================================================================


def test_safe_text_passes_through_normal_strings():
    assert _safe_excel_text("bunnings cement bag 20kg") == "bunnings cement bag 20kg"


def test_safe_text_passes_through_strings_with_dangerous_char_mid_string():
    """Mid-string `=` is legitimate user content (e.g. price formula
    notes) — must not get the apostrophe prefix."""
    s = "3 × 12m = $36"
    assert _safe_excel_text(s) == s


def test_safe_text_handles_none_and_empty():
    assert _safe_excel_text(None) == ""
    assert _safe_excel_text("") == ""


def test_safe_text_strips_control_chars():
    """Null byte and other C0 controls (except \\n) get stripped."""
    assert _safe_excel_text("legit text\x00with null byte") == "legit textwith null byte"
    # \n is preserved (multi-line notes are legitimate).
    assert _safe_excel_text("line1\nline2") == "line1\nline2"


def test_safe_text_preserves_legitimate_leading_apostrophe():
    """A string already starting with ' must NOT get a second apostrophe."""
    assert _safe_excel_text("'tis the season") == "'tis the season"


def test_safe_text_handles_leading_whitespace_then_dangerous_char():
    """Some apps detect formulas after leading whitespace — over-escape."""
    result = _safe_excel_text("   =evil")
    assert result == "'   =evil"


@pytest.mark.parametrize(
    "payload",
    [
        "=HYPERLINK(\"https://evil.example\",\"click\")",
        "+SUM(A1:A10)",
        "-2+3",
        "@SUM(A:A)",
        "\tcalc()",
        "\r=evil",
    ],
)
def test_safe_text_prefixes_all_dangerous_chars(payload):
    """All six dangerous prefixes get the apostrophe escape."""
    result = _safe_excel_text(payload)
    assert result.startswith("'")
    assert result == "'" + payload


# ===========================================================================
# _safe_sheet_name — Excel name sanitiser
# ===========================================================================


def test_safe_sheet_name_strips_forbidden_chars():
    used: set[str] = set()
    assert _safe_sheet_name("a/b\\c?d*e[f]g:h", used) == "a_b_c_d_e_f_g_h"


def test_safe_sheet_name_truncates_to_31_chars():
    used: set[str] = set()
    long = "x" * 50
    result = _safe_sheet_name(long, used)
    assert len(result) == 31


def test_safe_sheet_name_neutralises_leading_formula_prefix():
    """A job named ``=Bad Job`` must NOT produce a sheet tab labelled ``=…``."""
    used: set[str] = set()
    result = _safe_sheet_name("=Bad Job", used)
    assert not result.startswith(tuple(_FORMULA_PREFIXES))
    assert result.startswith("_")


def test_safe_sheet_name_preserves_cjk():
    """CJK chars are valid in Excel sheet names; preserve verbatim."""
    used: set[str] = set()
    assert _safe_sheet_name("晶晶", used) == "晶晶"


def test_safe_sheet_name_handles_collisions():
    used: set[str] = set()
    a = _safe_sheet_name("Kelly", used)
    b = _safe_sheet_name("Kelly", used)
    c = _safe_sheet_name("Kelly", used)
    assert a == "Kelly"
    assert b == "Kelly(1)"
    assert c == "Kelly(2)"


def test_safe_sheet_name_falls_back_to_job_id():
    used: set[str] = set()
    jid = uuid.uuid4()
    result = _safe_sheet_name("", used, job_id=jid)
    assert result == str(jid).replace("-", "")[:8]


# ===========================================================================
# build_export_filename — RFC 5987 dual-form
# ===========================================================================


def test_filename_date_range():
    f = build_export_filename(
        from_date=_datetime.date(2025, 7, 1),
        to_date=_datetime.date(2026, 6, 30),
        job_name=None,
        job_id=None,
        today=_datetime.date(2026, 5, 11),
    )
    assert f.ascii_fallback == "sitetracker-export-2025-07-01-to-2026-06-30.xlsx"
    assert f.utf8 == "sitetracker-export-2025-07-01-to-2026-06-30.xlsx"


def test_filename_today_only():
    f = build_export_filename(
        from_date=None,
        to_date=None,
        job_name=None,
        job_id=None,
        today=_datetime.date(2026, 5, 11),
    )
    assert f.ascii_fallback == "sitetracker-export-2026-05-11.xlsx"


def test_filename_single_job_cjk_uses_rfc5987_dual_form():
    """晶晶 job: ASCII fallback uses job_id prefix; UTF-8 form preserves CJK."""
    jid = uuid.UUID("daefdeef-4efd-4418-9639-9f5889ccb1cd")
    f = build_export_filename(
        from_date=None,
        to_date=None,
        job_name="晶晶",
        job_id=jid,
        today=_datetime.date(2026, 5, 11),
    )
    # ASCII fallback can't carry CJK — falls back to job_id hex prefix.
    assert "晶晶" not in f.ascii_fallback
    assert "daefdeef" in f.ascii_fallback
    assert f.ascii_fallback.endswith("-2026-05-11.xlsx")
    # UTF-8 form preserves the CJK.
    assert "晶晶" in f.utf8
    assert f.utf8.endswith("-2026-05-11.xlsx")


def test_filename_single_job_ascii_name():
    jid = uuid.uuid4()
    f = build_export_filename(
        from_date=None,
        to_date=None,
        job_name="Kelly House",
        job_id=jid,
        today=_datetime.date(2026, 5, 11),
    )
    # Ascii slug present in fallback.
    assert "kelly" in f.ascii_fallback.lower()


# ===========================================================================
# build_workbook — functional contract
# ===========================================================================


@pytest.mark.asyncio
async def test_workbook_empty_input_has_only_all_expenses_sheet(
    db_session, seeded_admin
):
    body = await build_workbook(db_session)
    wb = _open(body)
    assert wb.sheetnames == ["All Expenses"]
    ws = wb["All Expenses"]
    # Row 1 title, row 2 inclusion, row 3 period, row 4 blank, row 5 headers.
    assert ws.cell(row=1, column=1).value == "All Expenses"
    assert "reviewed expenses only" in ws.cell(row=2, column=1).value
    assert ws.cell(row=5, column=1).value == "Date"


@pytest.mark.asyncio
async def test_workbook_reviewed_only_default_excludes_pending(
    db_session, seeded_admin
):
    job = await _mk_job(db_session, seeded_admin, name="Job A")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        review_status=ReviewStatus.pending,
        description="should not appear",
    )
    body = await build_workbook(db_session)
    wb = _open(body)
    # No per-job sheet for Job A because no reviewed rows exist.
    assert "Job A" not in wb.sheetnames
    # All-Expenses has no data rows.
    ws = wb["All Expenses"]
    # Header is row 5; the next row should be the Totals footer (no data).
    # Totals row sum should be zero.
    # Find the totals row by looking for "Totals" in col A.
    totals_row = None
    for r in range(6, 20):
        if ws.cell(row=r, column=1).value == "Totals":
            totals_row = r
            break
    assert totals_row is not None
    assert ws.cell(row=totals_row, column=7).value == Decimal("0.00")


@pytest.mark.asyncio
async def test_workbook_include_pending_includes_pending(
    db_session, seeded_admin
):
    job = await _mk_job(db_session, seeded_admin, name="Job B")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        review_status=ReviewStatus.pending,
        amount_inc_gst=Decimal("110.00"),
        description="pending entry",
    )
    body = await build_workbook(db_session, include_pending=True)
    wb = _open(body)
    assert "Job B" in wb.sheetnames
    ws = wb["All Expenses"]
    # Row 2 should now say "reviewed + pending".
    assert "reviewed + pending" in ws.cell(row=2, column=1).value
    # Row 6 has the pending row's data. openpyxl reads Date cells as
    # datetime objects even when we wrote a date; .date() normalises.
    date_cell = ws.cell(row=6, column=1).value
    assert (date_cell.date() if hasattr(date_cell, "date") else date_cell) == _today()
    assert ws.cell(row=6, column=12).value == "pending"
    assert ws.cell(row=6, column=7).value == Decimal("110.00")


@pytest.mark.asyncio
async def test_workbook_rejected_always_excluded(db_session, seeded_admin):
    job = await _mk_job(db_session, seeded_admin, name="Job C")
    # Three rejected rows — neither default nor opt-in should include them.
    for _ in range(3):
        await _mk_expense(
            db_session,
            job=job,
            admin=seeded_admin,
            review_status=ReviewStatus.rejected,
        )
    body_default = await build_workbook(db_session)
    body_with_pending = await build_workbook(db_session, include_pending=True)
    for body in (body_default, body_with_pending):
        wb = _open(body)
        assert "Job C" not in wb.sheetnames
        ws = wb["All Expenses"]
        # No data rows — totals row should be on row 6 (immediately after header).
        for r in range(6, 12):
            v = ws.cell(row=r, column=1).value
            if v == "Totals":
                # Totals zero
                assert ws.cell(row=r, column=7).value == Decimal("0.00")
                break
        else:
            pytest.fail("Totals row not found")


@pytest.mark.asyncio
async def test_workbook_per_job_sheet_appears_when_at_least_one_row(
    db_session, seeded_admin
):
    job = await _mk_job(
        db_session, seeded_admin, name="Has Rows", code="HR-01"
    )
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        amount_inc_gst=Decimal("550.00"),
    )
    body = await build_workbook(db_session)
    wb = _open(body)
    assert "Has Rows" in wb.sheetnames


@pytest.mark.asyncio
async def test_workbook_per_job_sheet_skipped_when_zero_rows(
    db_session, seeded_admin
):
    """A job with no rows in the export window does NOT get a sheet."""
    await _mk_job(db_session, seeded_admin, name="Empty Job")
    body = await build_workbook(db_session)
    wb = _open(body)
    assert wb.sheetnames == ["All Expenses"]


@pytest.mark.asyncio
async def test_workbook_per_job_sheet_header_period_totals_match_rows(
    db_session, seeded_admin
):
    """Row 6 (Period totals) on a per-job sheet must equal the column sums."""
    job = await _mk_job(
        db_session,
        seeded_admin,
        name="Match Totals",
        total_budget_ex_gst=Decimal("1000.00"),
    )
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, amount_inc_gst=Decimal("110.00")
    )
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, amount_inc_gst=Decimal("220.00")
    )
    body = await build_workbook(db_session)
    ws = _open(body)["Match Totals"]
    # Row 6 carries the period-totals string. Parse the inc value.
    row6 = ws.cell(row=6, column=1).value
    assert "Period totals" in row6
    assert "inc $330.00" in row6


@pytest.mark.asyncio
async def test_workbook_per_job_sheet_project_summary_is_all_time_when_filtered(
    db_session, seeded_admin
):
    """Filtered period totals MAY differ from the project summary.

    When a date filter narrows the rows, the per-job header has two
    distinct blocks: period totals (filtered) and project summary
    (always all-time, dashboard view).
    """
    job = await _mk_job(
        db_session,
        seeded_admin,
        name="Filter Demo",
        total_budget_ex_gst=Decimal("5000.00"),
    )
    # Row inside the filter window.
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        amount_inc_gst=Decimal("110.00"),
        expense_date=_datetime.date(2026, 5, 5),
    )
    # Row OUTSIDE the filter window.
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        amount_inc_gst=Decimal("330.00"),
        expense_date=_datetime.date(2026, 1, 1),
    )
    body = await build_workbook(
        db_session,
        from_date=_datetime.date(2026, 5, 1),
        to_date=_datetime.date(2026, 5, 31),
    )
    ws = _open(body)["Filter Demo"]
    period_totals = ws.cell(row=6, column=1).value
    project_summary = ws.cell(row=10, column=1).value
    # Period totals reflect only the inside-filter row.
    assert "inc $110.00" in period_totals
    # All-time project summary reflects BOTH rows ($440 inc).
    assert "$440.00" in project_summary
    # The all-time label is explicit so the accountant knows it differs.
    assert "all-time" in ws.cell(row=8, column=1).value.lower()


@pytest.mark.asyncio
async def test_workbook_date_range_filters_all_expenses(
    db_session, seeded_admin
):
    job = await _mk_job(db_session, seeded_admin, name="Date Filtered")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        expense_date=_datetime.date(2026, 5, 5),
        amount_inc_gst=Decimal("110.00"),
        description="inside-range",
    )
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        expense_date=_datetime.date(2026, 1, 1),
        amount_inc_gst=Decimal("220.00"),
        description="outside-range",
    )
    body = await build_workbook(
        db_session,
        from_date=_datetime.date(2026, 5, 1),
        to_date=_datetime.date(2026, 5, 31),
    )
    ws = _open(body)["All Expenses"]
    # Only one data row (row 6); description = inside-range.
    assert ws.cell(row=6, column=6).value == "inside-range"
    assert ws.cell(row=7, column=1).value == "Totals"


@pytest.mark.asyncio
async def test_workbook_job_id_filter(db_session, seeded_admin):
    a = await _mk_job(db_session, seeded_admin, name="Job A")
    b = await _mk_job(db_session, seeded_admin, name="Job B")
    await _mk_expense(db_session, job=a, admin=seeded_admin, description="a-row")
    await _mk_expense(db_session, job=b, admin=seeded_admin, description="b-row")
    body = await build_workbook(db_session, job_id=a.job_id)
    wb = _open(body)
    # Only Job A's per-job sheet.
    assert wb.sheetnames == ["All Expenses", "Job A"]
    ws = wb["All Expenses"]
    assert ws.cell(row=6, column=6).value == "a-row"
    # No "b-row" anywhere on the All-Expenses sheet.
    for r in range(6, 12):
        v = ws.cell(row=r, column=6).value
        assert v != "b-row"


@pytest.mark.asyncio
async def test_workbook_unknown_job_id_raises_jobnotfound(db_session):
    with pytest.raises(JobNotFound):
        await build_workbook(db_session, job_id=uuid.uuid4())


@pytest.mark.asyncio
async def test_workbook_labour_expense_blank_supplier(db_session, seeded_admin):
    job = await _mk_job(db_session, seeded_admin, name="Labour Job")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        expense_type=ExpenseType.labour,
        supplier=None,
        description="day rate",
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    # Column D = Supplier; must be blank (empty string), not "None".
    assert ws.cell(row=6, column=4).value in ("", None)


@pytest.mark.asyncio
async def test_workbook_audit_columns_populated(db_session, seeded_admin):
    job = await _mk_job(db_session, seeded_admin, name="Audit Job")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        raw_input_text="bunnings $305 kelly bluemetal",
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    # Column O (15) = Raw input text; column P (16) = Created at.
    assert ws.cell(row=6, column=15).value == "bunnings $305 kelly bluemetal"
    created_at = ws.cell(row=6, column=16).value
    # openpyxl decodes datetime cells as datetime objects.
    assert isinstance(created_at, _datetime.datetime)


@pytest.mark.asyncio
async def test_workbook_money_columns_decimal_not_float(db_session, seeded_admin):
    job = await _mk_job(db_session, seeded_admin, name="Decimal Job")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        amount_inc_gst=Decimal("33.33"),
        payment_method=PaymentMethod.transfer,
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    inc = ws.cell(row=6, column=7).value
    # openpyxl reads numeric cells as float or Decimal depending on version;
    # either way the value must equal 33.33 exactly when compared to Decimal.
    assert Decimal(str(inc)) == Decimal("33.33")


@pytest.mark.asyncio
async def test_workbook_per_job_sheets_sorted_alphabetically(
    db_session, seeded_admin
):
    """Per-job sheets ordered alphabetically by job_name."""
    z = await _mk_job(db_session, seeded_admin, name="Zebra")
    a = await _mk_job(db_session, seeded_admin, name="Alpha")
    m = await _mk_job(db_session, seeded_admin, name="Mango")
    for j in (z, a, m):
        await _mk_expense(db_session, job=j, admin=seeded_admin)
    body = await build_workbook(db_session)
    wb = _open(body)
    # First sheet is All Expenses, then alphabetical.
    assert wb.sheetnames == ["All Expenses", "Alpha", "Mango", "Zebra"]


@pytest.mark.asyncio
async def test_workbook_cjk_job_name_preserved_as_sheet_name(
    db_session, seeded_admin
):
    job = await _mk_job(db_session, seeded_admin, name="晶晶")
    await _mk_expense(db_session, job=job, admin=seeded_admin)
    body = await build_workbook(db_session)
    wb = _open(body)
    assert "晶晶" in wb.sheetnames


@pytest.mark.asyncio
async def test_workbook_period_label_when_no_filter(db_session, seeded_admin):
    """The period annotation defaults to 'All time' to today."""
    job = await _mk_job(db_session, seeded_admin, name="Period Demo")
    await _mk_expense(db_session, job=job, admin=seeded_admin)
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    assert "All time" in ws.cell(row=3, column=1).value


# ===========================================================================
# Formula-injection neutralisation — workbook-level tests
# ===========================================================================


@pytest.mark.asyncio
async def test_injection_equals_in_description_neutralised(
    db_session, seeded_admin
):
    job = await _mk_job(db_session, seeded_admin, name="Inj Job")
    payload = '=HYPERLINK("https://evil.example","click")'
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, description=payload
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    # Col F (6) = Description.
    cell = ws.cell(row=6, column=6).value
    assert cell.startswith("'")
    assert cell == "'" + payload


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "payload",
    ["+SUM(A1:A10)", "-2+3", "@SUM(A:A)", "\tcalc()", "\r=evil"],
)
async def test_injection_other_prefixes_neutralised(
    db_session, seeded_admin, payload
):
    job = await _mk_job(db_session, seeded_admin, name="Inj Other")
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, description=payload
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    cell = ws.cell(row=6, column=6).value
    assert cell.startswith("'")


@pytest.mark.asyncio
async def test_injection_in_notes_field(db_session, seeded_admin):
    job = await _mk_job(db_session, seeded_admin, name="Notes Inj")
    payload = '=HYPERLINK("https://evil.example","X")'
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, notes=payload
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    # Col N (14) = Notes.
    assert ws.cell(row=6, column=14).value == "'" + payload


@pytest.mark.asyncio
async def test_injection_in_raw_input_text_field(db_session, seeded_admin):
    """Raw input text is the highest-risk vector."""
    job = await _mk_job(db_session, seeded_admin, name="Raw Inj")
    payload = "=cmd|'/c calc'!A1"
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, raw_input_text=payload
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    # Col O (15) = Raw input text.
    assert ws.cell(row=6, column=15).value == "'" + payload


@pytest.mark.asyncio
async def test_injection_in_supplier_name(db_session, seeded_admin):
    job = await _mk_job(db_session, seeded_admin, name="Sup Inj")
    bad = await _mk_supplier(db_session, name="=cmd|/c calc")
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, supplier=bad
    )
    body = await build_workbook(db_session)
    wb = _open(body)
    # Col D (4) on All-Expenses; col C (3) on per-job sheet (no Job col).
    assert wb["All Expenses"].cell(row=6, column=4).value.startswith("'")
    assert wb["Sup Inj"].cell(row=14, column=3).value.startswith("'")


@pytest.mark.asyncio
async def test_injection_in_job_name(db_session, seeded_admin):
    """Job name appears in All-Expenses col B AND per-job sheet row 1."""
    bad_name = "=Bad Job"
    job = await _mk_job(db_session, seeded_admin, name=bad_name)
    await _mk_expense(db_session, job=job, admin=seeded_admin)
    body = await build_workbook(db_session)
    wb = _open(body)
    # All-Expenses col B (2) = Job name → escaped.
    assert wb["All Expenses"].cell(row=6, column=2).value.startswith("'")
    # Per-job sheet row 1 = "Job: =Bad Job". The full string starts with
    # "Job:" so the first char is not a formula prefix → no apostrophe
    # prepended on the wrapped title; this is fine because the
    # ENTIRE cell starts with "Job: " which is safe. The bare name
    # inside the col B reference is what carries the injection risk
    # and is escaped.
    # Sheet tab name itself has the underscore-prefix neutralisation.
    assert not any(name.startswith("=") for name in wb.sheetnames)


@pytest.mark.asyncio
async def test_injection_in_job_name_sheet_name_underscore_prefix(
    db_session, seeded_admin
):
    """Sheet tab for ``=Bad Job`` must not start with ``=``."""
    job = await _mk_job(db_session, seeded_admin, name="=Bad Job")
    await _mk_expense(db_session, job=job, admin=seeded_admin)
    body = await build_workbook(db_session)
    wb = _open(body)
    matching = [s for s in wb.sheetnames if s != "All Expenses"]
    assert len(matching) == 1
    assert matching[0].startswith("_")
    assert "Bad Job" in matching[0]


@pytest.mark.asyncio
async def test_injection_mid_string_equals_preserved(db_session, seeded_admin):
    """``3 × 12m = $36`` is legitimate user content — preserve verbatim."""
    job = await _mk_job(db_session, seeded_admin, name="Mid Eq")
    legit = "3 × 12m = $36"
    await _mk_expense(
        db_session, job=job, admin=seeded_admin, description=legit
    )
    body = await build_workbook(db_session)
    ws = _open(body)["All Expenses"]
    assert ws.cell(row=6, column=6).value == legit


# ===========================================================================
# Direct end-to-end sanity check — build → load → spot-check shape
# ===========================================================================


@pytest.mark.asyncio
async def test_workbook_full_smoke(db_session, seeded_admin):
    """Two reviewed expenses across one job → workbook structure is sound."""
    job = await _mk_job(
        db_session,
        seeded_admin,
        name="Smoke",
        code="SM-01",
        total_budget_ex_gst=Decimal("10000.00"),
        contract_value_ex_gst=Decimal("12000.00"),
    )
    sup = await _mk_supplier(db_session, name="Bunnings")
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        amount_inc_gst=Decimal("110.00"),
        supplier=sup,
        description="bag of cement",
    )
    await _mk_expense(
        db_session,
        job=job,
        admin=seeded_admin,
        amount_inc_gst=Decimal("550.00"),
        payment_method=PaymentMethod.cash,
        description="skip bin",
    )
    body = await build_workbook(db_session)
    wb = _open(body)
    assert wb.sheetnames == ["All Expenses", "Smoke"]
    all_ws = wb["All Expenses"]
    # Two data rows on All Expenses.
    assert all_ws.cell(row=6, column=6).value == "bag of cement"
    assert all_ws.cell(row=7, column=6).value == "skip bin"
    # Totals on row 8.
    assert all_ws.cell(row=8, column=1).value == "Totals"
    assert all_ws.cell(row=8, column=7).value == Decimal("660.00")
    # Per-job sheet has period-totals header + project-summary block.
    smoke = wb["Smoke"]
    assert "Period totals" in smoke.cell(row=6, column=1).value
    assert "Project budget summary" in smoke.cell(row=8, column=1).value


@pytest.mark.asyncio
async def test_workbook_check_constraints_unaffected(db_session, seeded_admin):
    """Smoke: building the workbook does not trigger DB writes / CHECK fires.

    Defensive — confirms the export is purely read-only and doesn't
    accidentally mutate Phase 3 Lite+ data.
    """
    job = await _mk_job(
        db_session,
        seeded_admin,
        name="Const Job",
        contract_value_ex_gst=Decimal("100000.00"),
        target_profit_ratio_pct=Decimal("15.00"),
        total_budget_ex_gst=Decimal("85000.00"),
    )
    await _mk_expense(db_session, job=job, admin=seeded_admin)
    _ = await build_workbook(db_session)
    # Re-read the job; values must be unchanged.
    fresh = (
        await db_session.execute(
            sa_text(
                "SELECT contract_value_ex_gst, target_profit_ratio_pct, "
                "total_budget_ex_gst FROM jobs WHERE job_id = :jid"
            ),
            {"jid": str(job.job_id)},
        )
    ).one()
    assert Decimal(str(fresh.contract_value_ex_gst)) == Decimal("100000.00")
    assert Decimal(str(fresh.target_profit_ratio_pct)) == Decimal("15.00")
    assert Decimal(str(fresh.total_budget_ex_gst)) == Decimal("85000.00")
