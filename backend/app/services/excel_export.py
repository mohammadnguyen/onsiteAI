"""Phase 4 — Excel export / accountant handoff.

Pure read-only export over Phase 1 + 2 + 3 Lite data. Generates a single
``.xlsx`` workbook on demand with:

* an ``All Expenses`` sheet — every reviewed (default) or reviewed+pending
  (opt-in) expense across every job, BAS-friendly columns
* one per-job sheet for every job that has at least one row in the
  export window, with a date-range-aware two-block header

Frozen by ``docs/phase-4-plan.md``. Key contracts:

* **Default inclusion rule is reviewed-only.** Rejected always excluded.
  Pending opt-in via ``include_pending=True``. Intentionally stricter
  than Phase 3 Lite's dashboard ``(reviewed, pending)`` rule — the
  dashboard answers "how much could we owe (worst case)?", the
  accountant export answers "what is confirmed (defensible)?".
* **Per-job sheet headers are date-range aware.** When ``from_date`` /
  ``to_date`` is set, "all-time" project numbers are NOT presented as
  if they match the rows on the sheet. The header is split into two
  clearly-labelled blocks: an **Export period totals** block (computed
  from the rows actually shown, respecting the active filter +
  inclusion rule) and a **Project budget summary** block (always
  all-time, always Phase 3 Lite's reviewed+pending dashboard view,
  labelled explicitly).
* **Excel formula-injection protection.** Every text cell goes through
  :func:`_safe_excel_text`. Values starting with ``=``, ``+``, ``-``,
  ``@``, ``\\t``, or ``\\r`` (after leading whitespace) get an
  apostrophe prefix so the spreadsheet app renders them as inert text,
  not as formulas. CWE-1236 mitigation.
* **Sheet name sanitisation** also neutralises a leading formula
  prefix — a tab labelled ``=Bad Job`` is hostile UX even though sheet
  names don't evaluate formulas.

Decimal precision: money cells are written as native ``Decimal`` (via
openpyxl); ``date`` cells as ``date`` objects; ``Created at`` as
``datetime`` objects. Cell number-formats govern the visual display.
String cells are the only formula-injection vector and the only path
routed through :func:`_safe_excel_text`.
"""

from __future__ import annotations

import datetime as _datetime
import re
import uuid
from collections.abc import Sequence
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.expense import Expense, ReviewStatus
from app.models.job import Job
from app.services.budget_summary import summarize_job
from app.services.jobs import JobNotFound

# ---------------------------------------------------------------------------
# Formula-injection neutralisation
# ---------------------------------------------------------------------------

# Excel / LibreOffice / Numbers treat cells starting with one of these
# characters (after leading whitespace) as formulas. Frozen by
# docs/phase-4-plan.md — mirror this set in any future writer.
_FORMULA_PREFIXES = frozenset(("=", "+", "-", "@", "\t", "\r"))

# C0 control characters except \t (0x09), \n (0x0A), \r (0x0D). \t and
# \r are kept so leading-formula detection can see them; \n is kept
# because legitimate multi-line notes use it.
_STRIP_CONTROL = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")


def _safe_excel_text(value: str | None) -> str:
    """Return ``value`` neutralised for Excel formula-injection.

    Rules (frozen by ``docs/phase-4-plan.md``):

    * ``None`` / empty → ``""`` (no apostrophe spam).
    * Embedded C0 control chars except ``\\n`` are stripped.
    * Leading whitespace (ASCII space only — ``\\t`` and ``\\r`` are
      themselves dangerous, never legitimate "leading whitespace")
      is preserved in the output but the danger check is performed
      after stripping it.
    * If the first non-space char is one of ``=``, ``+``, ``-``,
      ``@``, ``\\t``, ``\\r``, prepend ``'`` (Excel's documented
      "force as text" prefix). The apostrophe is preserved in the
      stored cell value but not displayed by Excel.
    * Strings already starting with a legitimate apostrophe are NOT
      double-prefixed — only the dangerous prefixes get the escape.
    """
    if value is None:
        return ""
    cleaned = _STRIP_CONTROL.sub("", value)
    if cleaned == "":
        return ""
    # Skip leading ASCII space (only — \t/\r are themselves dangerous
    # leading chars; we don't treat them as "leading whitespace").
    i = 0
    while i < len(cleaned) and cleaned[i] == " ":
        i += 1
    if i >= len(cleaned):
        # Whitespace-only string — no danger.
        return cleaned
    if cleaned[i] in _FORMULA_PREFIXES:
        return "'" + cleaned
    return cleaned


# ---------------------------------------------------------------------------
# Sheet name sanitisation
# ---------------------------------------------------------------------------

# Excel forbids these in sheet names: \ / ? * [ ] :
_SHEET_FORBIDDEN = re.compile(r"[\\/?*\[\]:]")
_EXCEL_SHEET_NAME_MAX = 31


def _safe_sheet_name(
    name: str, used_names: set[str], job_id: uuid.UUID | None = None
) -> str:
    """Build a unique, Excel-valid sheet name from a job name.

    Algorithm (frozen):
      1. Strip / replace forbidden chars with ``_``.
      2. Truncate to 31 chars.
      3. If the result starts with a formula prefix, prepend ``_``
         (sheet names don't evaluate formulas in Excel, but a tab
         labelled ``=Bad Job`` is hostile UX and a clear injection
         signal worth neutralising).
      4. Empty-after-strip → use first 8 chars of ``job_id`` hex.
      5. If collides with an existing name in ``used_names``, suffix
         with ``(1)``, ``(2)``, … (truncating the base to fit).

    Mutates ``used_names`` by adding the returned name.
    """
    cleaned = _SHEET_FORBIDDEN.sub("_", (name or "").strip())
    cleaned = cleaned[:_EXCEL_SHEET_NAME_MAX]
    if cleaned and cleaned[0] in _FORMULA_PREFIXES:
        # Prepend underscore; truncate to keep within 31 chars.
        cleaned = ("_" + cleaned)[:_EXCEL_SHEET_NAME_MAX]
    if not cleaned:
        cleaned = (
            str(job_id).replace("-", "")[:8] if job_id else "Sheet"
        )

    base = cleaned
    counter = 1
    while cleaned in used_names:
        suffix = f"({counter})"
        cleaned = base[: _EXCEL_SHEET_NAME_MAX - len(suffix)] + suffix
        counter += 1
    used_names.add(cleaned)
    return cleaned


# ---------------------------------------------------------------------------
# Filename helper (RFC 5987 dual-form for non-ASCII job names)
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class ExportFilename:
    """Pair of filenames for the Content-Disposition header.

    ``ascii_fallback`` populates the plain ``filename="…"`` form (some
    older clients can't read RFC 5987). ``utf8`` is the un-encoded
    UTF-8 form that the endpoint percent-encodes for the
    ``filename*=UTF-8''…`` parameter.
    """

    ascii_fallback: str
    utf8: str


def build_export_filename(
    *,
    from_date: date | None,
    to_date: date | None,
    job_name: str | None,
    job_id: uuid.UUID | None,
    today: date,
) -> ExportFilename:
    """Compute the workbook filename per the frozen rules.

    * Date range set: ``sitetracker-export-{from}-to-{to}.xlsx``
    * Single-job filter: ``sitetracker-export-{slug-or-id}-{today}.xlsx``
    * Otherwise: ``sitetracker-export-{today}.xlsx``

    For non-ASCII job names (e.g. ``晶晶``) the UTF-8 form preserves
    the characters; the ASCII fallback falls back to the first 8 chars
    of ``job_id`` so a download still works on clients that don't
    honour RFC 5987.
    """
    if from_date and to_date:
        stamp = f"{from_date.isoformat()}-to-{to_date.isoformat()}"
        return ExportFilename(
            ascii_fallback=f"sitetracker-export-{stamp}.xlsx",
            utf8=f"sitetracker-export-{stamp}.xlsx",
        )
    if job_id is not None and job_name:
        ascii_slug = _ascii_slug(job_name)
        if not ascii_slug:
            ascii_slug = str(job_id).replace("-", "")[:8]
        return ExportFilename(
            ascii_fallback=f"sitetracker-export-{ascii_slug}-{today.isoformat()}.xlsx",
            utf8=f"sitetracker-export-{_safe_utf8_slug(job_name)}-{today.isoformat()}.xlsx",
        )
    stamp = today.isoformat()
    return ExportFilename(
        ascii_fallback=f"sitetracker-export-{stamp}.xlsx",
        utf8=f"sitetracker-export-{stamp}.xlsx",
    )


_NON_ASCII_OR_PUNCT = re.compile(r"[^a-zA-Z0-9]+")


def _ascii_slug(name: str) -> str:
    """Return an ASCII-safe slug, or "" if nothing survives.

    Strips formula-prefix chars too — defence-in-depth so a job named
    ``=evil`` can't sneak its prefix into the ``filename=`` form.
    """
    # Drop everything that isn't ASCII alphanumeric.
    cleaned = _NON_ASCII_OR_PUNCT.sub("-", name).strip("-").lower()
    # Drop leading formula prefix if any survived the alphanumeric filter
    # (defence-in-depth — the regex above should already have stripped them).
    while cleaned and cleaned[0] in _FORMULA_PREFIXES:
        cleaned = cleaned[1:]
    return cleaned[:60]


def _safe_utf8_slug(name: str) -> str:
    """Sanitise a UTF-8 slug for use in the ``filename*=UTF-8''…`` form.

    Replace whitespace and Excel-forbidden chars but preserve CJK
    chars (``晶晶``) verbatim. The endpoint percent-encodes the result.
    """
    # Same forbidden chars as sheet names + spaces.
    cleaned = re.sub(r"[\s\\/?*\[\]:]+", "-", name).strip("-")
    while cleaned and cleaned[0] in _FORMULA_PREFIXES:
        cleaned = cleaned[1:]
    return cleaned[:60]


# ---------------------------------------------------------------------------
# Workbook builder
# ---------------------------------------------------------------------------

_ALL_EXPENSES_HEADERS: Sequence[str] = (
    "Date",
    "Job",
    "Job code",
    "Supplier",
    "Category",
    "Description",
    "Amount inc GST",
    "GST amount",
    "Amount ex GST",
    "Payment method",
    "Receipt status",
    "Review status",
    "Entered by",
    "Notes",
    "Raw input text",
    "Created at",
    "Expense ID",
)

_PER_JOB_HEADERS: Sequence[str] = (
    # Same as All Expenses, minus "Job" (every row on the sheet is the
    # same job — printing the column would be noise).
    "Date",
    "Job code",
    "Supplier",
    "Category",
    "Description",
    "Amount inc GST",
    "GST amount",
    "Amount ex GST",
    "Payment method",
    "Receipt status",
    "Review status",
    "Entered by",
    "Notes",
    "Raw input text",
    "Created at",
    "Expense ID",
)

_MONEY_FORMAT = '"$"#,##0.00'
_DATE_FORMAT = "DD/MM/YYYY"
# ISO-UTC for Created at — unambiguous; locale conversion deferred per plan.
_DATETIME_FORMAT = 'YYYY-MM-DD"T"HH:MM:SS"Z"'


def _inclusion_label(include_pending: bool) -> str:
    return (
        "Inclusion rule: reviewed + pending"
        if include_pending
        else "Inclusion rule: reviewed expenses only"
    )


def _period_label(
    from_date: date | None, to_date: date | None, today: date
) -> str:
    fd = from_date.isoformat() if from_date else "All time"
    td = to_date.isoformat() if to_date else today.isoformat()
    return f"Export period: {fd} to {td}"


async def fetch_export_expenses(
    db: AsyncSession,
    *,
    from_date: date | None,
    to_date: date | None,
    job_id: uuid.UUID | None,
    include_pending: bool,
) -> list[Expense]:
    """Load the expenses inside the active inclusion rule + filters.

    Public because the PDF report service consumes the SAME frozen
    inclusion rule — duplicating it would let the two accountant-facing
    exports drift apart.

    Inclusion rule (frozen, distinct from Phase 3 Lite):

    * ``reviewed`` — always included
    * ``pending`` — included only when ``include_pending=True``
    * ``rejected`` — always excluded
    """
    statuses: list[ReviewStatus] = [ReviewStatus.reviewed]
    if include_pending:
        statuses.append(ReviewStatus.pending)

    q = (
        select(Expense)
        .options(
            selectinload(Expense.job),
            # supplier, category, entered_by are lazy="joined" on the model.
        )
        .where(Expense.review_status.in_(statuses))
        # Total order for a REPRODUCIBLE export: expense_date + created_at can
        # tie (multiple same-day captures share the transaction timestamp), so
        # add expense_id as a final tiebreak — otherwise row order falls to
        # Postgres heap order and the same export can differ run-to-run /
        # host-to-host.
        .order_by(
            Expense.expense_date.asc(),
            Expense.created_at.asc(),
            Expense.expense_id.asc(),
        )
    )
    if from_date is not None:
        q = q.where(Expense.expense_date >= from_date)
    if to_date is not None:
        q = q.where(Expense.expense_date <= to_date)
    if job_id is not None:
        q = q.where(Expense.job_id == job_id)
    return list((await db.execute(q)).scalars().all())


def _to_naive_utc(dt: datetime | None) -> datetime | None:
    """Strip timezone from a tz-aware datetime, converting to UTC first.

    openpyxl rejects tz-aware datetimes (Excel itself has no timezone
    concept). Our ``expenses.created_at`` is ``DateTime(timezone=True)``
    storing UTC values; we present them in the workbook as naive UTC
    timestamps and label the cell format with a trailing ``Z`` so the
    accountant can see they're UTC.
    """
    if dt is None:
        return None
    if dt.tzinfo is None:
        return dt
    return dt.astimezone(tz=_datetime.UTC).replace(tzinfo=None)


def _row_for_expense(
    e: Expense, *, include_job_columns: bool
) -> list[object]:
    """Build a single data row for a worksheet.

    Money values stay as Decimal so openpyxl writes them as numeric.
    Dates stay as ``date`` objects. ``Created at`` is converted to a
    naive UTC datetime (openpyxl rejects tz-aware datetimes). Every
    text value routes through :func:`_safe_excel_text` for formula-
    injection neutralisation.

    Per-job sheets drop ONLY the ``Job`` (name) column — ``Job code``
    is preserved on both sheets because one job can have one code and
    it remains a useful reference for the accountant. The row layout
    therefore matches ``_ALL_EXPENSES_HEADERS`` minus ``Job`` for the
    per-job case.
    """
    # Pre-compute all string fields through the safety helper so the
    # row construction below is just an ordered list.
    job_name = _safe_excel_text(e.job.job_name if e.job else None)
    job_code = _safe_excel_text(e.job.job_code if e.job else None)
    supplier = _safe_excel_text(e.supplier.supplier_name if e.supplier else None)
    category = _safe_excel_text(
        e.category.category_name if e.category else None
    )
    description = _safe_excel_text(e.description)
    payment = _safe_excel_text(e.payment_method.value)
    receipt = _safe_excel_text(e.receipt_status.value)
    review = _safe_excel_text(e.review_status.value)
    entered_by = _safe_excel_text(e.entered_by.email if e.entered_by else None)
    notes = _safe_excel_text(e.notes)
    raw_input = _safe_excel_text(e.raw_input_text)
    created_at = _to_naive_utc(e.created_at)
    expense_id = _safe_excel_text(str(e.expense_id))

    tail: list[object] = [
        supplier,
        category,
        description,
        e.amount_inc_gst,
        e.gst_amount,
        e.amount_ex_gst,
        payment,
        receipt,
        review,
        entered_by,
        notes,
        raw_input,
        created_at,
        expense_id,
    ]
    if include_job_columns:
        # All-Expenses sheet: Date, Job, Job code, then the common tail.
        return [e.expense_date, job_name, job_code, *tail]
    # Per-job sheet: Date, Job code, then the common tail (Job dropped).
    return [e.expense_date, job_code, *tail]


def _set_number_formats(
    ws,
    *,
    money_cols: Sequence[int],
    date_col: int,
    created_at_col: int,
    last_row: int,
) -> None:
    """Apply Excel number-format strings to money / date / datetime columns."""
    for r in range(2, last_row + 1):
        for col in money_cols:
            ws.cell(row=r, column=col).number_format = _MONEY_FORMAT
        ws.cell(row=r, column=date_col).number_format = _DATE_FORMAT
        ws.cell(row=r, column=created_at_col).number_format = _DATETIME_FORMAT


def _build_all_expenses_sheet(
    wb: Workbook,
    expenses: list[Expense],
    *,
    inclusion_label: str,
    period_label: str,
) -> None:
    """Build the cross-job ``All Expenses`` sheet.

    Annotation rows 1–3 + blank row 4 + header on row 5 + data + totals.
    """
    ws = wb.active
    ws.title = "All Expenses"

    bold = Font(bold=True)

    # Row 1: title
    ws.cell(row=1, column=1, value=_safe_excel_text("All Expenses")).font = Font(bold=True, size=14)
    # Row 2: inclusion rule
    ws.cell(row=2, column=1, value=_safe_excel_text(inclusion_label))
    # Row 3: period
    ws.cell(row=3, column=1, value=_safe_excel_text(period_label))
    # Row 4: blank (skip)
    # Row 5: column headers
    header_row = 5
    for col_idx, header in enumerate(_ALL_EXPENSES_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=_safe_excel_text(header))
        cell.font = bold
    # Freeze panes below the header row + after the first (Date) column.
    ws.freeze_panes = "B6"

    # Data rows starting at row 6
    first_data_row = header_row + 1
    for offset, e in enumerate(expenses):
        row = first_data_row + offset
        values = _row_for_expense(e, include_job_columns=True)
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=value)
    last_data_row = first_data_row + len(expenses) - 1 if expenses else first_data_row - 1

    # Cell number-formats for money / date / created_at columns
    # All-Expenses col indices:
    # 1=Date, 2=Job, 3=Job code, 4=Supplier, 5=Category, 6=Description,
    # 7=inc, 8=GST, 9=ex, 10=Payment, 11=Receipt, 12=Review, 13=Entered by,
    # 14=Notes, 15=Raw input, 16=Created at, 17=Expense ID
    if last_data_row >= first_data_row:
        _set_number_formats(
            ws,
            money_cols=(7, 8, 9),
            date_col=1,
            created_at_col=16,
            last_row=last_data_row,
        )

    # Footer row: Totals
    totals_row = last_data_row + 1 if expenses else first_data_row
    ws.cell(row=totals_row, column=1, value=_safe_excel_text("Totals")).font = bold
    inc_total = sum((e.amount_inc_gst for e in expenses), Decimal("0.00"))
    gst_total = sum((e.gst_amount for e in expenses), Decimal("0.00"))
    ex_total = sum((e.amount_ex_gst for e in expenses), Decimal("0.00"))
    ws.cell(row=totals_row, column=7, value=inc_total).number_format = _MONEY_FORMAT
    ws.cell(row=totals_row, column=8, value=gst_total).number_format = _MONEY_FORMAT
    ws.cell(row=totals_row, column=9, value=ex_total).number_format = _MONEY_FORMAT
    # Bold the totals money cells too.
    for c in (7, 8, 9):
        ws.cell(row=totals_row, column=c).font = bold


async def _build_job_sheet(
    wb: Workbook,
    db: AsyncSession,
    job: Job,
    expenses: list[Expense],
    *,
    sheet_name: str,
    inclusion_label: str,
    period_label: str,
) -> None:
    """Build a per-job sheet with the two-block header.

    Row layout (frozen by docs/phase-4-plan.md):

      Row 1:  Job: {job_name}
      Row 2:  Job code: ... · Site: ...
      Row 3:  (blank)
      Row 4:  Export period: ...
      Row 5:  Inclusion rule: ...
      Row 6:  Period totals (these rows): inc / GST / ex
      Row 7:  (blank)
      Row 8:  Project budget summary (all-time, dashboard view ...):
      Row 9:  Contract value ex GST: ... · Total budget ex GST: ...
      Row 10: All-time spent inc GST / ex GST / GST
      Row 11: Remaining ex GST / % consumed (all-time)
      Row 12: (blank)
      Row 13: column headers
      Row 14+: data
      Row N+1: footer "Totals"
    """
    ws = wb.create_sheet(title=sheet_name)
    bold = Font(bold=True)
    italic_muted = Font(italic=True, color="606060")

    # Row 1: title
    _title_cell = ws.cell(row=1, column=1, value=_safe_excel_text(f"Job: {job.job_name}"))
    _title_cell.font = Font(bold=True, size=14)
    # Row 2: code + site
    code = job.job_code or "—"
    site = job.site_address or "—"
    ws.cell(
        row=2,
        column=1,
        value=_safe_excel_text(f"Job code: {code}  ·  Site: {site}"),
    )
    # Row 3 blank.
    # Row 4: period label (italic muted)
    ws.cell(row=4, column=1, value=_safe_excel_text(period_label)).font = italic_muted
    # Row 5: inclusion rule (italic muted)
    ws.cell(row=5, column=1, value=_safe_excel_text(inclusion_label)).font = italic_muted
    # Row 6: period totals (these rows)
    inc_total = sum((e.amount_inc_gst for e in expenses), Decimal("0.00"))
    gst_total = sum((e.gst_amount for e in expenses), Decimal("0.00"))
    ex_total = sum((e.amount_ex_gst for e in expenses), Decimal("0.00"))
    ws.cell(
        row=6,
        column=1,
        value=_safe_excel_text(
            f"Period totals (these rows): inc ${inc_total:,.2f}  ·  "
            f"GST ${gst_total:,.2f}  ·  ex ${ex_total:,.2f}"
        ),
    ).font = bold
    # Row 7 blank.
    # Row 8: project summary label
    ws.cell(
        row=8,
        column=1,
        value=_safe_excel_text(
            "Project budget summary (all-time, dashboard view — "
            "may differ from period totals above):"
        ),
    ).font = italic_muted
    # Rows 9-11: dashboard summary
    summary = await summarize_job(db, job.job_id)
    contract = (
        f"${summary.total_budget_ex_gst:,.2f}"
        if summary.total_budget_ex_gst is not None
        else "—"
    )
    contract_val = (
        f"${job.contract_value_ex_gst:,.2f}"
        if job.contract_value_ex_gst is not None
        else "—"
    )
    ws.cell(
        row=9,
        column=1,
        value=_safe_excel_text(
            f"Contract value ex GST: {contract_val}  ·  "
            f"Total budget ex GST: {contract}"
        ),
    )
    ws.cell(
        row=10,
        column=1,
        value=_safe_excel_text(
            f"All-time spent inc GST: ${summary.actual_inc_gst:,.2f}  ·  "
            f"ex GST: ${summary.actual_ex_gst:,.2f}  ·  "
            f"GST: ${summary.gst_amount:,.2f}"
        ),
    )
    remaining = (
        f"${summary.remaining_ex_gst:,.2f}"
        if summary.remaining_ex_gst is not None
        else "—"
    )
    percent = (
        f"{summary.percent_consumed:.2f}%"
        if summary.percent_consumed is not None
        else "—"
    )
    ws.cell(
        row=11,
        column=1,
        value=_safe_excel_text(
            f"Remaining ex GST: {remaining}  ·  "
            f"% consumed (all-time): {percent}"
        ),
    )
    # Row 12 blank.
    # Row 13: column headers
    header_row = 13
    for col_idx, header in enumerate(_PER_JOB_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=_safe_excel_text(header))
        cell.font = bold
    ws.freeze_panes = f"B{header_row + 1}"

    # Data rows
    first_data_row = header_row + 1
    for offset, e in enumerate(expenses):
        row = first_data_row + offset
        values = _row_for_expense(e, include_job_columns=False)
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=row, column=col_idx, value=value)
    last_data_row = first_data_row + len(expenses) - 1 if expenses else first_data_row - 1

    # Per-job col indices (no Job col):
    # 1=Date, 2=Job code, 3=Supplier, 4=Category, 5=Description,
    # 6=inc, 7=GST, 8=ex, 9=Payment, 10=Receipt, 11=Review,
    # 12=Entered by, 13=Notes, 14=Raw input, 15=Created at, 16=Expense ID
    if last_data_row >= first_data_row:
        _set_number_formats(
            ws,
            money_cols=(6, 7, 8),
            date_col=1,
            created_at_col=15,
            last_row=last_data_row,
        )

    # Footer row: Totals — matches the period-totals header row.
    totals_row = last_data_row + 1 if expenses else first_data_row
    ws.cell(row=totals_row, column=1, value=_safe_excel_text("Totals")).font = bold
    ws.cell(row=totals_row, column=6, value=inc_total).number_format = _MONEY_FORMAT
    ws.cell(row=totals_row, column=7, value=gst_total).number_format = _MONEY_FORMAT
    ws.cell(row=totals_row, column=8, value=ex_total).number_format = _MONEY_FORMAT
    for c in (6, 7, 8):
        ws.cell(row=totals_row, column=c).font = bold


async def build_workbook(
    db: AsyncSession,
    *,
    from_date: date | None = None,
    to_date: date | None = None,
    job_id: uuid.UUID | None = None,
    include_pending: bool = False,
    today: date | None = None,
) -> bytes:
    """Compose the `.xlsx` workbook for the accountant export.

    See module docstring for the frozen contract. Returns the raw bytes
    of the workbook so the API layer can stream it with the appropriate
    ``Content-Type`` + ``Content-Disposition``.

    Raises :class:`JobNotFound` when ``job_id`` is supplied but doesn't
    resolve to a persisted job.
    """
    if today is None:
        today = date.today()

    # Validate job_id resolves (so 404 fires before we waste a workbook build).
    if job_id is not None:
        job = (
            await db.execute(select(Job).where(Job.job_id == job_id))
        ).scalar_one_or_none()
        if job is None:
            raise JobNotFound(job_id)

    expenses = await fetch_export_expenses(
        db,
        from_date=from_date,
        to_date=to_date,
        job_id=job_id,
        include_pending=include_pending,
    )

    inclusion_label = _inclusion_label(include_pending)
    period_label = _period_label(from_date, to_date, today)

    wb = Workbook()
    _build_all_expenses_sheet(
        wb,
        expenses,
        inclusion_label=inclusion_label,
        period_label=period_label,
    )

    # Per-job sheets — one per job that has at least one row in the
    # export window. Sort alphabetically by job_name.
    by_job: dict[uuid.UUID, list[Expense]] = {}
    job_lookup: dict[uuid.UUID, Job] = {}
    for e in expenses:
        by_job.setdefault(e.job_id, []).append(e)
        job_lookup.setdefault(e.job_id, e.job)
    used_names: set[str] = {wb.active.title}
    for jid in sorted(by_job.keys(), key=lambda j: job_lookup[j].job_name.lower()):
        job_obj = job_lookup[jid]
        sheet_name = _safe_sheet_name(job_obj.job_name, used_names, job_id=jid)
        await _build_job_sheet(
            wb,
            db,
            job_obj,
            by_job[jid],
            sheet_name=sheet_name,
            inclusion_label=inclusion_label,
            period_label=period_label,
        )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
